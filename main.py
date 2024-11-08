import os
import argparse
import subprocess
import time
import laspy
import openpyxl

def convert_zlas_to_las(in_zlas, target_folder):
    """
    Converts a ZLAS file to a LAS file using ArcGIS Pro's Python environment.
    """
    # Path to ArcGIS Pro Python executable
    python_executable = r"C:\Program Files\ArcGIS\Pro\bin\Python\envs\arcgispro-py3\python.exe"
    
    # Path to the arcpy conversion script
    script_path = r"conversion.py"  # This is a new script that does the arcpy conversion
    
    # Command to run the arcpy script using ArcGIS Pro Python
    command = [python_executable, script_path, in_zlas, target_folder]
    
    # Run the conversion process
    process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    
    if process.returncode != 0:
        print(f"Error running the conversion on {in_zlas}: {stderr.decode()}")
    else:
        print(f"Successfully converted {in_zlas} to LAS.")

def delete_converted_files(las_file, lasx_file):
    """
    Deletes the converted LAS or LASX file if it exists.
    """
    # Delete LAS file if it exists
    if os.path.exists(las_file):
        os.remove(las_file)
        print(f"Deleted the converted LAS file: {las_file}")
    
    # Delete LASX file if it exists
    if os.path.exists(lasx_file):
        os.remove(lasx_file)
        print(f"Deleted the converted LASX file: {lasx_file}")

def read_las_header_to_excel(file_path, sheet, row_index):
    """
    Reads the header of a LAS file and writes the information to an Excel sheet.
    """
    with laspy.open(file_path) as las_file:
        header = las_file.header

        # Write header information to the Excel sheet
        sheet.cell(row=row_index, column=1).value = header.file_source_id
        sheet.cell(row=row_index, column=2).value = str(header.global_encoding)  # Convert to string
        sheet.cell(row=row_index, column=3).value = f"{header.version.major}.{header.version.minor}"  # Format version
        sheet.cell(row=row_index, column=4).value = header.system_identifier
        sheet.cell(row=row_index, column=5).value = header.generating_software
        sheet.cell(row=row_index, column=6).value = header.creation_date
        sheet.cell(row=row_index, column=7).value = header.offset_to_point_data
        sheet.cell(row=row_index, column=8).value = str(header.point_format)  # Convert point_format to string
        sheet.cell(row=row_index, column=9).value = header.point_count

        # Convert scales, offsets, maxs, and mins to strings
        sheet.cell(row=row_index, column=10).value = ', '.join(map(str, header.scales))
        sheet.cell(row=row_index, column=11).value = ', '.join(map(str, header.offsets))
        sheet.cell(row=row_index, column=12).value = ', '.join(map(str, header.maxs))
        sheet.cell(row=row_index, column=13).value = ', '.join(map(str, header.mins))

def process_zlas_files_in_folder(folder_path, sheet, wb, report_path):
    """
    Processes all .zlas files in the specified folder.
    """
    # List all .zlas files in the folder
    zlas_files = [f for f in os.listdir(folder_path) if f.endswith('.zlas')]
    
    for zlas_file in zlas_files:
        in_zlas = os.path.join(folder_path, zlas_file)
        
        # Convert ZLAS to LAS using ArcGIS Pro Python
        convert_zlas_to_las(in_zlas, folder_path)
        
        # Assuming the conversion generates a LAS file with the same name as the ZLAS file but with a .las extension
        las_file = os.path.join(folder_path, zlas_file.replace('.zlas', '.las'))
        
        # Assuming the conversion might also generate a .lasx file with the same name
        lasx_file = os.path.join(folder_path, zlas_file.replace('.zlas', '.lasx'))
        
        # Process the LAS file header and save it to Excel
        next_row = sheet.max_row + 1
        read_las_header_to_excel(las_file, sheet, next_row)

        # Save workbook incrementally after each file is processed
        wb.save(report_path)
        
        # Delete the converted LAS and LASX files after processing
        delete_converted_files(las_file, lasx_file)

def main():
    # Set up argument parsing
    parser = argparse.ArgumentParser(description='Convert ZLAS file to LAS file, process header and save to Excel.')
    parser.add_argument('folder', type=str, help='Path to the folder containing .zlas files')
    parser.add_argument('--report', type=str, default='output.xlsx', 
                        help='Path to the output Excel report file (default: output.xlsx)')

    args = parser.parse_args()

    folder_path = args.folder
    report_path = args.report

    # Ensure the file path is absolute
    folder_path = os.path.abspath(folder_path)

    # Create or open the workbook to store results
    if os.path.exists(report_path):
        wb = openpyxl.load_workbook(report_path)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "LAS Header Report"
        
        # Write header for Excel columns
        headers = [
            "File Source ID", "Global Encoding", "Version", "System Identifier",
            "Generating Software", "Creation Date", "Offset to Point Data",
            "Point Data Format", "Number of Point Records", "Scale Factor",
            "Offset", "Max X, Y, Z", "Min X, Y, Z"
        ]
        sheet.append(headers)

    # Process all .zlas files in the folder and pass the sheet and workbook to the function
    process_zlas_files_in_folder(folder_path, sheet, wb, report_path)

    # Indicate completion
    print("Done")

if __name__ == "__main__":
    main()
